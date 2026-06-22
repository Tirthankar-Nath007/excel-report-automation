from __future__ import annotations

import io
import json

import openpyxl
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter

from app.models.report import AggregatedSummary

BLOCK_COL_OFFSET = 11  # each portfolio block starts 11 cols after the previous


class ExcelExportService:
    def generate(
        self, aggregated_by_portfolio: dict[str, AggregatedSummary]
    ) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        portfolios = sorted(aggregated_by_portfolio.keys())
        for idx, portfolio in enumerate(portfolios):
            col_start = 1 + idx * BLOCK_COL_OFFSET
            self._write_portfolio_block(
                ws, aggregated_by_portfolio[portfolio], col_start
            )

        self._auto_width(ws)

        # Hidden metadata sheet for merge service
        ds = wb.create_sheet("_data")
        raw: dict[str, dict] = {
            p: {
                "from_date": s.from_date,
                "to_date": s.to_date,
                "data_source_summary": s.data_source_summary,
                "error_summary": s.error_summary,
            }
            for p, s in aggregated_by_portfolio.items()
        }
        ds["A1"] = json.dumps(raw)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    def _write_portfolio_block(
        self, ws: openpyxl.worksheet.worksheet.Worksheet,
        summary: AggregatedSummary,
        col_start: int,
    ) -> None:
        row = 1

        # Section title
        title_cell = ws.cell(row=row, column=col_start, value=summary.portfolio.upper())
        title_cell.font = Font(bold=True, size=12)
        row += 2

        # Data Source summary
        row = self._write_table(
            ws, row, col_start,
            headers=("Type", "Count"),
            rows=sorted(summary.data_source_summary.items()),
            grand_total=sum(summary.data_source_summary.values()),
        )
        row += 2

        # Error summary
        self._write_table(
            ws, row, col_start,
            headers=("Status", "Count"),
            rows=sorted(summary.error_summary.items()),
            grand_total=sum(summary.error_summary.values()),
        )

    def _write_table(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        col_start: int,
        headers: tuple[str, str],
        rows: list[tuple[str, int]],
        grand_total: int,
    ) -> int:
        bold = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font = bold
            c.border = border
        row += 1

        # Data rows
        for label, count in rows:
            lc = ws.cell(row=row, column=col_start, value=label)
            lc.border = border
            cc = ws.cell(row=row, column=col_start + 1, value=count)
            cc.border = border
            row += 1

        # Grand Total
        gt_label = ws.cell(row=row, column=col_start, value="Grand Total")
        gt_label.font = bold
        gt_label.border = border
        gt_count = ws.cell(row=row, column=col_start + 1, value=grand_total)
        gt_count.font = bold
        gt_count.border = border

        return row + 1

    def _auto_width(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        col_widths: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_widths[cell.column] = max(
                        col_widths.get(cell.column, 0),
                        len(str(cell.value)) + 2,
                    )
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(width, 60)
